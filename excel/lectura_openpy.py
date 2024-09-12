import openpyxl

# Cargar el archivo Excel
archivo_excel = "./datos.xlsx"
libro_trabajo = openpyxl.load_workbook(archivo_excel)

# Acceder a una hoja específica
hoja = libro_trabajo["Hoja1"]

# Recorrer las filas de la hoja de cálculo
for fila in hoja.iter_rows(min_row=1, values_only=True):
    print(fila)

# Recorrer las columnas de la hoja de cálculo
for columna in hoja.iter_cols(values_only=True):
    print(columna)

# Recorrer las celdas de la hoja de cálculo
for fila in hoja.iter_rows(min_row=1, values_only=True):
    for celda in fila:
        print(celda)

# Insertar una nueva columna después de la columna B (por ejemplo)
hoja.insert_cols(2)

# Insertar una nueva fila después de la fila 2 (por ejemplo)
hoja.insert_rows(2)

# Insertar un nuevo valor en la celda B2 (por ejemplo)
hoja['B2'] = "Nuevo Valor"

# Guardar el archivo Excel con la nueva columna
libro_trabajo.save(archivo_excel)

# Cerrar el archivo Excel
libro_trabajo.close()
